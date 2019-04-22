from openpyxl import load_workbook
wb = load_workbook('固定资产电子台账1.xlsx')
sheet = wb.get_sheet_by_name('卡片列表')
print(sheet.max_row)
print(sheet.max_column)



b4_too = sheet.cell(row=3, column=4)
print(b4_too.value)